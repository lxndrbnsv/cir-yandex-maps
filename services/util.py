import re
import os
import shutil
from openpyxl import load_workbook


class ClearWB:
    def __init__(self):
        os.remove("./results.xlsx")
        shutil.copy("./template.xlsx", "./results.xlsx")


class CheckWB:
    def __init__(self):
        wb = load_workbook(filename="results.xlsx")
        ws = wb["Worksheet"]

        print(ws.max_row)


class NormalizePhoneNumber:
    def __init__(self, phone_string):
        stripped_symbols = (
            phone_string.replace("(", "")
            .replace(")", "")
            .replace(" ", "")
            .replace("-", "")
        )
        if stripped_symbols[0] == "8":
            stripped_list = list(stripped_symbols)
            stripped_list[0] = "+7"
            normalized = "".join(stripped_list)
        else:
            normalized = stripped_symbols

        self.normalize = normalized


def parse_query_string(query: str) -> str:
    address_items = query.split(",")

    address_data = [address_items[0]]

    for a in address_items[1:]:
        digits = re.findall(r"\d+", a)
        try:
            address_data.append(digits[0])
        except IndexError:
            pass

    return f"#{'-'.join(address_data)}"
